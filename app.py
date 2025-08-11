import streamlit as st
import json
import os
from typing import List, Dict, Any, Optional, Tuple
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Fill, Border, Alignment
import google.generativeai as genai
import tempfile
import logging
from io import BytesIO
import pandas as pd
import math
import re
from collections import Counter, defaultdict
from datetime import datetime
import numpy as np

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Page config
st.set_page_config(
    page_title="📊 AI Excel Editor - Advanced",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

class AdvancedExcelProcessor:
    """Advanced Excel processor that handles formulas, merged cells, and complex structures"""
    
    def __init__(self):
        self.workbooks = {}
        self.file_paths = {}
        self.sheet_structures = {}
        self.search_index = {}
        self.merged_cells_info = {}
        
    def load_excel_files(self, uploaded_files) -> Dict[str, Any]:
        """Load Excel files with advanced processing"""
        results = {}
        
        for uploaded_file in uploaded_files:
            try:
                # Save to temporary file
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                    tmp_file.write(uploaded_file.read())
                    tmp_path = tmp_file.name
                
                self.file_paths[uploaded_file.name] = tmp_path
                
                # Load workbook with all data and formulas
                workbook = load_workbook(tmp_path, data_only=False)
                self.workbooks[uploaded_file.name] = workbook
                
                # Process all sheets
                file_data = {
                    'file_name': uploaded_file.name,
                    'sheets': {},
                    'total_sheets': len(workbook.sheetnames)
                }
                
                for sheet_name in workbook.sheetnames:
                    sheet_data = self._process_sheet_advanced(workbook[sheet_name], uploaded_file.name, sheet_name)
                    file_data['sheets'][sheet_name] = sheet_data
                
                results[uploaded_file.name] = file_data
                
            except Exception as e:
                logger.error(f"Error loading {uploaded_file.name}: {str(e)}")
                results[uploaded_file.name] = {'error': str(e)}
        
        # Build search index after all files loaded
        self._build_advanced_search_index()
        
        return results
    
    def _process_sheet_advanced(self, sheet, file_name: str, sheet_name: str) -> Dict[str, Any]:
        """Advanced sheet processing with formula and formatting handling"""
        try:
            if not sheet.max_row or sheet.max_row == 0:
                return {
                    'grid_data': [],
                    'merged_cells': [],
                    'formulas': {},
                    'max_row': 0,
                    'max_col': 0,
                    'headers': []
                }
            
            max_row = sheet.max_row
            max_col = sheet.max_column or 0
            
            # Store merged cells info
            merged_ranges = []
            for merged_range in sheet.merged_cells.ranges:
                merged_ranges.append({
                    'range': str(merged_range),
                    'start_row': merged_range.min_row,
                    'end_row': merged_range.max_row,
                    'start_col': merged_range.min_col,
                    'end_col': merged_range.max_col
                })
            
            # Build complete grid data
            grid_data = []
            formulas = {}
            search_content = []
            
            for row_num in range(1, max_row + 1):
                row_data = []
                row_search_text = []
                
                for col_num in range(1, max_col + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    cell_ref = f"{get_column_letter(col_num)}{row_num}"
                    
                    # Get cell value (display value)
                    display_value = ""
                    if cell.value is not None:
                        display_value = str(cell.value)
                    
                    # Check for formulas
                    if cell.data_type == 'f':  # Formula
                        formulas[cell_ref] = cell.value
                        # Try to get calculated value
                        try:
                            calc_workbook = load_workbook(self.file_paths[file_name], data_only=True)
                            calc_cell = calc_workbook[sheet_name].cell(row=row_num, column=col_num)
                            if calc_cell.value is not None:
                                display_value = str(calc_cell.value)
                        except:
                            pass
                    
                    # Store cell info
                    cell_info = {
                        'value': display_value,
                        'formula': cell.value if cell.data_type == 'f' else None,
                        'row': row_num,
                        'col': col_num,
                        'ref': cell_ref,
                        'is_merged': self._is_cell_merged(cell_ref, merged_ranges)
                    }
                    
                    row_data.append(cell_info)
                    
                    # Add to search content if has meaningful value
                    if display_value and len(display_value.strip()) > 0:
                        row_search_text.append(f"{get_column_letter(col_num)}: {display_value}")
                
                grid_data.append(row_data)
                
                # Add to search index
                if row_search_text:
                    search_content.append({
                        'file_name': file_name,
                        'sheet_name': sheet_name,
                        'row_number': row_num,
                        'content': " | ".join(row_search_text),
                        'row_data': {cell['ref']: cell['value'] for cell in row_data if cell['value']}
                    })
            
            # Extract headers (typically from first few rows)
            headers = []
            if grid_data:
                for cell in grid_data[0]:  # First row
                    if cell['value']:
                        headers.append(cell['value'])
                    else:
                        headers.append(f"Col {get_column_letter(cell['col'])}")
            
            # Store for search
            sheet_key = f"{file_name}:{sheet_name}"
            self.search_index[sheet_key] = search_content
            self.merged_cells_info[sheet_key] = merged_ranges
            
            return {
                'grid_data': grid_data,
                'merged_cells': merged_ranges,
                'formulas': formulas,
                'max_row': max_row,
                'max_col': max_col,
                'headers': headers,
                'search_content': search_content
            }
            
        except Exception as e:
            logger.error(f"Error processing sheet {sheet_name}: {str(e)}")
            return {
                'grid_data': [],
                'merged_cells': [],
                'formulas': {},
                'max_row': 0,
                'max_col': 0,
                'headers': [],
                'error': str(e)
            }
    
    def _is_cell_merged(self, cell_ref: str, merged_ranges: List[Dict]) -> bool:
        """Check if a cell is part of a merged range"""
        try:
            from openpyxl.utils import coordinate_from_string
            col_letter, row_num = coordinate_from_string(cell_ref)
            col_num = column_index_from_string(col_letter)
            
            for merged_range in merged_ranges:
                if (merged_range['start_row'] <= row_num <= merged_range['end_row'] and
                    merged_range['start_col'] <= col_num <= merged_range['end_col']):
                    return True
            return False
        except:
            return False
    
    def _build_advanced_search_index(self):
        """Build TF-IDF search index from all processed data"""
        all_documents = []
        self.document_metadata = []
        
        for sheet_key, search_content in self.search_index.items():
            file_name, sheet_name = sheet_key.split(':', 1)
            
            for row_info in search_content:
                all_documents.append(row_info['content'])
                self.document_metadata.append({
                    'file_name': file_name,
                    'sheet_name': sheet_name,
                    'row_number': row_info['row_number'],
                    'row_data': row_info['row_data'],
                    'content': row_info['content']
                })
        
        # Simple TF-IDF implementation
        self.tfidf_index = SimpleTFIDF()
        if all_documents:
            self.tfidf_index.fit(all_documents)
    
    def search_data(self, query: str, top_k: int = 10) -> List[Dict[str, Any]]:
        """Advanced search using TF-IDF"""
        if not hasattr(self, 'tfidf_index') or not self.document_metadata:
            return []
        
        scores = self.tfidf_index.query(query)
        
        # Get top results
        top_indices = np.argsort(scores)[::-1][:top_k]
        
        results = []
        for idx in top_indices:
            if scores[idx] > 0:
                metadata = self.document_metadata[idx].copy()
                metadata['similarity_score'] = scores[idx]
                results.append(metadata)
        
        return results
    
    def read_cell(self, file_name: str, sheet_name: str, cell_ref: str) -> Dict[str, Any]:
        """Read specific cell by Excel reference (e.g., 'A1', 'B5')"""
        try:
            if file_name not in self.workbooks:
                return {'success': False, 'error': f"File {file_name} not found"}
            
            workbook = self.workbooks[file_name]
            if sheet_name not in workbook.sheetnames:
                return {'success': False, 'error': f"Sheet {sheet_name} not found"}
            
            sheet = workbook[sheet_name]
            
            # Parse cell reference
            from openpyxl.utils import coordinate_from_string
            col_letter, row_num = coordinate_from_string(cell_ref.upper())
            col_num = column_index_from_string(col_letter)
            
            if row_num > sheet.max_row or col_num > sheet.max_column:
                return {'success': False, 'error': f"Cell {cell_ref} is out of bounds"}
            
            cell = sheet.cell(row=row_num, column=col_num)
            
            # Get both formula and value
            cell_value = cell.value if cell.value is not None else ""
            is_formula = cell.data_type == 'f'
            
            # Get calculated value if it's a formula
            calculated_value = cell_value
            if is_formula:
                try:
                    calc_workbook = load_workbook(self.file_paths[file_name], data_only=True)
                    calc_cell = calc_workbook[sheet_name].cell(row=row_num, column=col_num)
                    calculated_value = calc_cell.value if calc_cell.value is not None else ""
                except:
                    pass
            
            return {
                'success': True,
                'cell_ref': cell_ref.upper(),
                'raw_value': str(cell_value),
                'calculated_value': str(calculated_value),
                'is_formula': is_formula,
                'location': f"{file_name}/{sheet_name}/{cell_ref.upper()}"
            }
            
        except Exception as e:
            return {'success': False, 'error': f"Error reading cell: {str(e)}"}
    
    def edit_cell(self, file_name: str, sheet_name: str, cell_ref: str, new_value: str) -> Dict[str, Any]:
        """Edit specific cell by Excel reference"""
        try:
            if file_name not in self.workbooks:
                return {'success': False, 'error': f"File {file_name} not found"}
            
            workbook = self.workbooks[file_name]
            if sheet_name not in workbook.sheetnames:
                return {'success': False, 'error': f"Sheet {sheet_name} not found"}
            
            sheet = workbook[sheet_name]
            
            # Parse cell reference
            from openpyxl.utils import coordinate_from_string
            col_letter, row_num = coordinate_from_string(cell_ref.upper())
            col_num = column_index_from_string(col_letter)
            
            if row_num > sheet.max_row or col_num > sheet.max_column:
                return {'success': False, 'error': f"Cell {cell_ref} is out of bounds"}
            
            cell = sheet.cell(row=row_num, column=col_num)
            old_value = cell.value if cell.value is not None else ""
            
            # Convert value to appropriate type
            try:
                # Try to detect if it should be a number
                if new_value.replace('.', '').replace('-', '').replace('+', '').isdigit():
                    new_value = float(new_value) if '.' in new_value else int(new_value)
                elif new_value.startswith('='):
                    # It's a formula, keep as string but Excel will interpret it
                    pass
            except:
                pass  # Keep as string
            
            # Update cell
            cell.value = new_value
            
            # Save workbook
            workbook.save(self.file_paths[file_name])
            
            # Refresh the sheet data
            sheet_data = self._process_sheet_advanced(sheet, file_name, sheet_name)
            
            return {
                'success': True,
                'cell_ref': cell_ref.upper(),
                'old_value': str(old_value),
                'new_value': str(new_value),
                'location': f"{file_name}/{sheet_name}/{cell_ref.upper()}"
            }
            
        except Exception as e:
            return {'success': False, 'error': f"Error editing cell: {str(e)}"}
    
    def get_sheet_dataframe(self, file_name: str, sheet_name: str) -> pd.DataFrame:
        """Convert sheet data to DataFrame for display"""
        try:
            sheet_key = f"{file_name}:{sheet_name}"
            if sheet_key not in self.search_index:
                return pd.DataFrame()
            
            # Get grid data
            if file_name not in self.workbooks:
                return pd.DataFrame()
            
            workbook = self.workbooks[file_name]
            sheet = workbook[sheet_name]
            
            # Create DataFrame from sheet
            data = []
            headers = []
            
            # Get headers from first row
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=1, column=col)
                header = cell.value if cell.value else f"Column_{get_column_letter(col)}"
                headers.append(str(header))
            
            # Get data rows
            for row in range(2, sheet.max_row + 1):
                row_data = []
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    value = cell.value if cell.value is not None else ""
                    row_data.append(str(value))
                data.append(row_data)
            
            if data:
                df = pd.DataFrame(data, columns=headers)
                return df
            else:
                return pd.DataFrame()
                
        except Exception as e:
            logger.error(f"Error creating DataFrame: {str(e)}")
            return pd.DataFrame()

class SimpleTFIDF:
    """Simple TF-IDF implementation for search"""
    
    def __init__(self):
        self.vocabulary = {}
        self.idf_values = {}
        self.documents = []
    
    def _tokenize(self, text: str) -> List[str]:
        """Simple tokenization"""
        text = re.sub(r'[^\w\s]', ' ', text.lower())
        tokens = text.split()
        stop_words = {'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of', 'with', 'by'}
        return [token for token in tokens if token not in stop_words and len(token) > 1]
    
    def fit(self, documents: List[str]):
        """Fit TF-IDF on documents"""
        self.documents = documents
        all_tokens = set()
        
        # Build vocabulary
        for doc in documents:
            tokens = self._tokenize(doc)
            all_tokens.update(tokens)
        
        self.vocabulary = {token: idx for idx, token in enumerate(all_tokens)}
        
        # Calculate IDF
        doc_count = len(documents)
        for token in self.vocabulary:
            docs_with_token = sum(1 for doc in documents if token in self._tokenize(doc))
            self.idf_values[token] = math.log(doc_count / (docs_with_token + 1))
    
    def query(self, query_text: str) -> np.ndarray:
        """Get similarity scores for query"""
        query_tokens = self._tokenize(query_text)
        scores = np.zeros(len(self.documents))
        
        for i, doc in enumerate(self.documents):
            doc_tokens = self._tokenize(doc)
            doc_token_counts = Counter(doc_tokens)
            
            score = 0
            for token in query_tokens:
                if token in self.vocabulary and token in doc_token_counts:
                    tf = doc_token_counts[token] / len(doc_tokens)
                    idf = self.idf_values.get(token, 0)
                    score += tf * idf
            
            scores[i] = score
        
        return scores

class AIExcelInstructor:
    """AI instructor that generates precise Excel instructions"""
    
    def __init__(self, api_key: str, excel_processor: AdvancedExcelProcessor):
        genai.configure(api_key=api_key)
        self.model = "gemini-2.0-flash-exp"
        self.excel_processor = excel_processor
        self.edit_history = []
    
    def process_request(self, user_request: str, file_structure: Dict[str, Any]) -> str:
        """Process user request and execute Excel operations"""
        try:
            # Create instruction prompt
            prompt = f"""
You are an expert Excel AI. You can search, read, and edit Excel files with precision.

AVAILABLE FILES:
{json.dumps(file_structure, indent=2)}

AVAILABLE COMMANDS:
1. SEARCH: [query] - Search for data across all files
2. READ_CELL: [filename] [sheetname] [cellref] - Read cell (e.g., READ_CELL: file.xlsx Sheet1 A5)
3. EDIT_CELL: [filename] [sheetname] [cellref] [newvalue] - Edit cell (e.g., EDIT_CELL: file.xlsx Sheet1 A5 "New Value")

CRITICAL RULES:
- Always use exact Excel cell references like A1, B5, C10, etc.
- Search first to find the exact location of data
- Use the search results to determine precise cell references
- File names may have spaces - handle them correctly

USER REQUEST: {user_request}

Execute the request step by step:
1. Search to find relevant data
2. Read specific cells if needed
3. Edit cells with exact references
4. Provide clear results

Format your response with actual command execution:
"""

            model = genai.GenerativeModel(self.model)
            response = model.generate_content(prompt)
            
            # Parse and execute commands from response
            commands = self._extract_commands(response.text)
            
            execution_results = ""
            if commands:
                execution_results = "\n🔧 COMMAND EXECUTION:\n"
                
                for i, command in enumerate(commands, 1):
                    execution_results += f"\n📋 Command {i}: {command}\n"
                    
                    result = self._execute_command(command)
                    execution_results += f"✅ Result: {self._format_result(result)}\n"
            
            # Generate final analysis
            final_prompt = f"""
Based on the execution results:
{execution_results}

Provide a comprehensive summary of:
1. What was accomplished
2. Specific changes made (with exact cell references)
3. Current state of the data
4. Any recommendations

Original request: {user_request}
"""
            
            final_response = model.generate_content(final_prompt)
            
            return f"""
🎯 AI EXCEL OPERATIONS:

{response.text}

{execution_results}

📊 COMPREHENSIVE SUMMARY:
{final_response.text}
"""
            
        except Exception as e:
            return f"Error processing request: {str(e)}"
    
    def _extract_commands(self, text: str) -> List[str]:
        """Extract executable commands from AI response"""
        commands = []
        lines = text.split('\n')
        
        for line in lines:
            line = line.strip()
            if any(cmd in line.upper() for cmd in ['SEARCH:', 'READ_CELL:', 'EDIT_CELL:']):
                # Clean up the command
                command = re.sub(r'^\d+\.\s*', '', line)  # Remove numbering
                commands.append(command.strip())
        
        return commands[:10]  # Limit to 10 commands
    
    def _execute_command(self, command: str) -> Dict[str, Any]:
        """Execute a single command"""
        try:
            command = command.strip()
            
            if command.upper().startswith('SEARCH:'):
                query = command[7:].strip()
                results = self.excel_processor.search_data(query, top_k=5)
                return {
                    'success': True,
                    'type': 'search',
                    'results': results,
                    'count': len(results)
                }
            
            elif command.upper().startswith('READ_CELL:'):
                parts = command[10:].strip().split()
                if len(parts) >= 3:
                    file_name = parts[0]
                    sheet_name = parts[1]
                    cell_ref = parts[2]
                    return self.excel_processor.read_cell(file_name, sheet_name, cell_ref)
                else:
                    return {'success': False, 'error': 'Invalid READ_CELL format'}
            
            elif command.upper().startswith('EDIT_CELL:'):
                # Parse EDIT_CELL: filename sheetname cellref newvalue
                parts = command[10:].strip().split(None, 3)
                if len(parts) >= 4:
                    file_name = parts[0]
                    sheet_name = parts[1]
                    cell_ref = parts[2]
                    new_value = parts[3].strip('"\'')  # Remove quotes if present
                    
                    result = self.excel_processor.edit_cell(file_name, sheet_name, cell_ref, new_value)
                    
                    if result['success']:
                        # Add to edit history
                        self.edit_history.append({
                            'timestamp': datetime.now().isoformat(),
                            'command': command,
                            'result': result
                        })
                    
                    return result
                else:
                    return {'success': False, 'error': 'Invalid EDIT_CELL format'}
            
            else:
                return {'success': False, 'error': f'Unknown command: {command}'}
                
        except Exception as e:
            return {'success': False, 'error': f'Error executing command: {str(e)}'}
    
    def _format_result(self, result: Dict[str, Any]) -> str:
        """Format command results for display"""
        if not result.get('success', False):
            return f"❌ {result.get('error', 'Unknown error')}"
        
        result_type = result.get('type', 'unknown')
        
        if result_type == 'search':
            if result['count'] > 0:
                formatted = f"Found {result['count']} results:\n"
                for res in result['results'][:3]:
                    formatted += f"   📍 {res['file_name']}/{res['sheet_name']}/Row {res['row_number']}\n"
                    formatted += f"   📊 {res['content'][:100]}...\n"
                return formatted
            else:
                return "No results found"
        
        elif 'cell_ref' in result:
            if 'old_value' in result:  # Edit result
                return f"✏️ Edited {result['cell_ref']}: '{result['old_value']}' → '{result['new_value']}'"
            else:  # Read result
                return f"📖 {result['cell_ref']}: {result.get('calculated_value', result.get('raw_value', ''))}"
        
        else:
            return f"✅ {str(result)[:100]}..."

# API Key
API_KEY = "AIzaSyCSDx-q3PgkvMQktdi4tScbT1wOLgZ9jQg"

# Initialize session state
if 'excel_processor' not in st.session_state:
    st.session_state.excel_processor = AdvancedExcelProcessor()
if 'ai_instructor' not in st.session_state:
    st.session_state.ai_instructor = None
if 'files_loaded' not in st.session_state:
    st.session_state.files_loaded = False
if 'file_structure' not in st.session_state:
    st.session_state.file_structure = {}
if 'analysis_result' not in st.session_state:
    st.session_state.analysis_result = ""

def create_excel_like_viewer(file_name: str, sheet_name: str, excel_processor: AdvancedExcelProcessor):
    """Create Excel-like viewer that actually shows the spreadsheet data"""
    try:
        if file_name not in excel_processor.workbooks:
            st.error("File not found")
            return
        
        workbook = excel_processor.workbooks[file_name]
        sheet = workbook[sheet_name]
        
        st.subheader(f"📊 {file_name} / {sheet_name}")
        
        # Show basic info
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Total Rows", sheet.max_row)
        col2.metric("Total Columns", sheet.max_column)
        
        # Get recent changes for highlighting
        changed_cells = set()
        if hasattr(st.session_state, 'ai_instructor') and st.session_state.ai_instructor:
            for edit in st.session_state.ai_instructor.edit_history:
                if file_name in edit['result'].get('location', ''):
                    location_parts = edit['result']['location'].split('/')
                    if len(location_parts) >= 3:
                        cell_ref = location_parts[-1]
                        changed_cells.add(cell_ref)
        
        col3.metric("🔥 Changed Cells", len(changed_cells))
        
        with col4:
            if st.button("🔄 Refresh View"):
                st.rerun()
        
        # Show change indicators
        if changed_cells:
            st.success(f"🔥 **Recently changed**: {', '.join(sorted(changed_cells))}")
        
        # Create the actual Excel-like table
        max_rows_to_show = 40
        max_cols_to_show = 12
        
        # Prepare data for display
        excel_data = []
        
        # Create header row with column letters
        header_row = [""] + [get_column_letter(col) for col in range(1, min(sheet.max_column + 1, max_cols_to_show + 1))]
        
        # Add data rows
        for row_num in range(1, min(sheet.max_row + 1, max_rows_to_show + 1)):
            row_data = [str(row_num)]  # Row number
            
            for col_num in range(1, min(sheet.max_column + 1, max_cols_to_show + 1)):
                cell = sheet.cell(row=row_num, column=col_num)
                cell_ref = f"{get_column_letter(col_num)}{row_num}"
                
                # Get cell value
                cell_value = ""
                if cell.value is not None:
                    if cell.data_type == 'f':  # Formula
                        try:
                            # Try to get calculated value
                            calc_workbook = load_workbook(excel_processor.file_paths[file_name], data_only=True)
                            calc_sheet = calc_workbook[sheet_name]
                            calc_cell = calc_sheet.cell(row=row_num, column=col_num)
                            if calc_cell.value is not None:
                                cell_value = str(calc_cell.value)
                            else:
                                cell_value = str(cell.value)  # Show formula if calc fails
                        except:
                            cell_value = str(cell.value)
                    else:
                        cell_value = str(cell.value)
                
                # Highlight changed cells
                if cell_ref in changed_cells:
                    cell_value = f"🔥 {cell_value}"
                
                row_data.append(cell_value)
            
            excel_data.append(row_data)
        
        # Create DataFrame with proper structure
        df = pd.DataFrame(excel_data, columns=header_row)
        
        # Display the Excel-like table
        st.dataframe(
            df,
            use_container_width=True,
            height=800,  # Taller for better viewing
            hide_index=True,
            column_config={
                "": st.column_config.TextColumn(
                    "",
                    width="small",
                    help="Row numbers"
                )
            }
        )
        
        # Quick navigation and tools
        st.divider()
        
        # Show specific cell info
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.write("**📍 Quick Cell Lookup:**")
            lookup_cell = st.text_input("Enter cell (e.g., A32):", key=f"lookup_{file_name}_{sheet_name}")
            if lookup_cell and st.button("🔍 Find Cell", key=f"find_{file_name}_{sheet_name}"):
                try:
                    from openpyxl.utils import coordinate_from_string
                    col_letter, row_num = coordinate_from_string(lookup_cell.upper())
                    cell = sheet.cell(row=row_num, column=column_index_from_string(col_letter))
                    
                    if cell.value is not None:
                        if cell.data_type == 'f':
                            st.info(f"**{lookup_cell.upper()}**: Formula = `{cell.value}`")
                            # Try to show calculated value
                            try:
                                calc_workbook = load_workbook(excel_processor.file_paths[file_name], data_only=True)
                                calc_cell = calc_workbook[sheet_name].cell(row=row_num, column=column_index_from_string(col_letter))
                                if calc_cell.value is not None:
                                    st.success(f"**Calculated Value**: {calc_cell.value}")
                            except:
                                pass
                        else:
                            st.success(f"**{lookup_cell.upper()}**: {cell.value}")
                    else:
                        st.warning(f"**{lookup_cell.upper()}**: Empty cell")
                except Exception as e:
                    st.error(f"Invalid cell reference: {lookup_cell}")
        
        with col2:
            st.write("**📂 Download Options:**")
            # Download as CSV
            df_clean = df.copy()
            # Remove emojis for clean export
            for col in df_clean.columns:
                if col != "":
                    df_clean[col] = df_clean[col].astype(str).str.replace('🔥 ', '', regex=False)
            
            csv = df_clean.to_csv(index=False)
            st.download_button(
                "📥 Download CSV",
                data=csv,
                file_name=f"{file_name}_{sheet_name}.csv",
                mime="text/csv",
                key=f"csv_{file_name}_{sheet_name}"
            )
            
            # Download Excel
            try:
                file_path = excel_processor.file_paths[file_name]
                with open(file_path, 'rb') as f:
                    excel_data = f.read()
                
                st.download_button(
                    "📥 Download Excel",
                    data=excel_data,
                    file_name=f"modified_{file_name}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"excel_{file_name}_{sheet_name}"
                )
            except Exception as e:
                st.error(f"Download error: {str(e)}")
        
        with col3:
            st.write("**🔬 Sheet Analysis:**")
            
            # Count formulas
            formula_count = 0
            for row in range(1, min(sheet.max_row + 1, 100)):
                for col in range(1, min(sheet.max_column + 1, 20)):
                    cell = sheet.cell(row=row, column=col)
                    if cell.data_type == 'f':
                        formula_count += 1
            
            st.metric("Formulas Found", formula_count)
            
            # Show merge info
            sheet_key = f"{file_name}:{sheet_name}"
            if sheet_key in excel_processor.merged_cells_info:
                merged_count = len(excel_processor.merged_cells_info[sheet_key])
                st.metric("Merged Ranges", merged_count)
        
        # Detailed change history
        if changed_cells:
            with st.expander(f"📝 Detailed Edit History ({len(changed_cells)} changes)"):
                if hasattr(st.session_state, 'ai_instructor') and st.session_state.ai_instructor:
                    sheet_edits = [edit for edit in st.session_state.ai_instructor.edit_history 
                                  if file_name in edit['result'].get('location', '') and sheet_name in edit['result'].get('location', '')]
                    
                    for i, edit in enumerate(sheet_edits, 1):
                        st.write(f"**Change {i}:** {edit['timestamp'][:19]}")
                        st.write(f"📍 **Cell**: {edit['result']['location'].split('/')[-1]}")
                        st.write(f"🔄 **Change**: `{edit['result'].get('old_value', 'empty')}` → `{edit['result'].get('new_value', 'empty')}`")
                        st.divider()
        
        # Show sample formulas
        formulas_found = []
        for row in range(1, min(sheet.max_row + 1, 50)):
            for col in range(1, min(sheet.max_column + 1, 15)):
                cell = sheet.cell(row=row, column=col)
                if cell.data_type == 'f':
                    cell_ref = f"{get_column_letter(col)}{row}"
                    formulas_found.append((cell_ref, str(cell.value)))
                    if len(formulas_found) >= 10:
                        break
            if len(formulas_found) >= 10:
                break
        
        if formulas_found:
            with st.expander(f"🔬 Formulas in Sheet ({len(formulas_found)} shown)"):
                for cell_ref, formula in formulas_found:
                    st.code(f"{cell_ref}: {formula}")
        
    except Exception as e:
        st.error(f"Error displaying sheet: {str(e)}")
        st.exception(e)

def main():
    st.title("📊 Advanced AI Excel Editor")
    st.markdown("**AI understands your requests and precisely edits Excel files with formula support!**")
    
    # Sidebar
    with st.sidebar:
        st.header("🔧 System Status")
        
        if st.session_state.files_loaded:
            file_count = len(st.session_state.excel_processor.workbooks)
            st.success(f"✅ {file_count} files loaded")
            
            # Show search index status
            total_searchable = sum(len(content) for content in st.session_state.excel_processor.search_index.values())
            st.info(f"🔍 {total_searchable} rows indexed")
            
            # Show edit history
            if st.session_state.ai_instructor:
                edit_count = len(st.session_state.ai_instructor.edit_history)
                st.info(f"✏️ {edit_count} edits made")
        else:
            st.warning("⚠️ No files loaded")
        
        st.divider()
        
        # Quick operations
        st.header("🚀 Quick Operations")
        
        if st.session_state.files_loaded:
            # Manual search
            st.subheader("🔍 Manual Search")
            search_query = st.text_input("Search query:", placeholder="Total Liabilities")
            if st.button("🔍 Search") and search_query:
                results = st.session_state.excel_processor.search_data(search_query)
                if results:
                    st.success(f"Found {len(results)} results")
                    for result in results[:3]:
                        st.text(f"📍 {result['file_name']}/{result['sheet_name']}/Row {result['row_number']}")
                        st.text(f"📊 {result['content'][:100]}...")
                else:
                    st.info("No results found")
            
            # Manual cell operations
            st.subheader("📱 Manual Cell Operations")
            
            # File and sheet selector
            file_names = list(st.session_state.excel_processor.workbooks.keys())
            selected_file = st.selectbox("File:", file_names, key="manual_file")
            
            if selected_file:
                workbook = st.session_state.excel_processor.workbooks[selected_file]
                sheet_names = workbook.sheetnames
                selected_sheet = st.selectbox("Sheet:", sheet_names, key="manual_sheet")
                
                if selected_sheet:
                    cell_ref = st.text_input("Cell (e.g., A5):", placeholder="A5")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        if st.button("📖 Read Cell") and cell_ref:
                            result = st.session_state.excel_processor.read_cell(selected_file, selected_sheet, cell_ref)
                            if result['success']:
                                st.success(f"✅ {cell_ref}: {result['calculated_value']}")
                                if result['is_formula']:
                                    st.info(f"Formula: {result['raw_value']}")
                            else:
                                st.error(result['error'])
                    
                    with col2:
                        new_value = st.text_input("New value:", placeholder="New Value")
                        if st.button("✏️ Edit Cell") and cell_ref and new_value:
                            result = st.session_state.excel_processor.edit_cell(selected_file, selected_sheet, cell_ref, new_value)
                            if result['success']:
                                st.success(f"✅ Edited {cell_ref}")
                                st.text(f"'{result['old_value']}' → '{result['new_value']}'")
                            else:
                                st.error(result['error'])
            
            # Save files
            st.divider()
            if st.button("💾 Save All Changes", type="primary"):
                try:
                    saved_count = 0
                    for file_name, workbook in st.session_state.excel_processor.workbooks.items():
                        file_path = st.session_state.excel_processor.file_paths[file_name]
                        workbook.save(file_path)
                        saved_count += 1
                    st.success(f"✅ Saved {saved_count} files")
                except Exception as e:
                    st.error(f"❌ Error saving: {str(e)}")
        
        # Clear all
        if st.button("🗑️ Clear All", type="secondary"):
            for key in ['files_loaded', 'file_structure', 'analysis_result']:
                if key in st.session_state:
                    st.session_state[key] = False if key == 'files_loaded' else {}
            st.session_state.excel_processor = AdvancedExcelProcessor()
            st.session_state.ai_instructor = None
            st.rerun()
    
    # Main content
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.header("📁 Upload Excel Files")
        uploaded_files = st.file_uploader(
            "Upload Excel files for AI analysis and editing",
            type=['xlsx', 'xls'],
            accept_multiple_files=True,
            help="Upload Excel files with formulas, merged cells, and complex formatting"
        )
        
        if uploaded_files:
            st.success(f"📄 {len(uploaded_files)} file(s) ready")
            
            if st.button("🔄 Load & Process Files", type="primary"):
                with st.spinner("🔄 Loading and processing Excel files with advanced features..."):
                    try:
                        # Load files
                        file_structure = st.session_state.excel_processor.load_excel_files(uploaded_files)
                        st.session_state.file_structure = file_structure
                        
                        # Initialize AI instructor
                        st.session_state.ai_instructor = AIExcelInstructor(API_KEY, st.session_state.excel_processor)
                        st.session_state.files_loaded = True
                        
                        st.success("✅ Files loaded with advanced processing!")
                        
                        # Show summary
                        total_sheets = sum(data.get('total_sheets', 0) for data in file_structure.values() if 'total_sheets' in data)
                        st.info(f"📊 Processed {len(file_structure)} files with {total_sheets} sheets")
                        
                    except Exception as e:
                        st.error(f"❌ Error loading files: {str(e)}")
    
    with col2:
        st.header("🤖 AI Request")
        user_request = st.text_area(
            "Tell AI what you want to do with your Excel files:",
            placeholder="""Examples:
• Find "Total Liabilities & Shareholders' equity" and change it to "Total Liabilities"
• Search for all cells containing "revenue" and show their values
• Edit cell B5 in Balance Sheet to show "Updated Value"
• Find the highest value in column C and tell me its location
• Replace all instances of "Aug-22" with "August 2022"
• Calculate the sum of all values in row 15""",
            height=150,
            help="Use natural language - AI will understand and execute precise Excel operations"
        )
        
        if st.button("🎯 Execute AI Request", type="primary", disabled=not (st.session_state.files_loaded and user_request)):
            with st.spinner("🤖 AI is analyzing your request and executing Excel operations..."):
                try:
                    result = st.session_state.ai_instructor.process_request(user_request, st.session_state.file_structure)
                    st.session_state.analysis_result = result
                    st.success("✅ AI request completed!")
                    
                    # Force rerun to show updated Excel viewer
                    st.rerun()
                    
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")
    
    # Excel File Viewer - Show IMMEDIATELY after files are loaded
    if st.session_state.files_loaded:
        st.divider()
        st.header("📊 Live Excel Viewer")
        st.markdown("**View your Excel files as they are being edited in real-time**")
        
        # Auto-select first file and sheet for immediate viewing
        file_names = list(st.session_state.excel_processor.workbooks.keys())
        if file_names:
            # Create columns for file/sheet selection
            col1, col2, col3 = st.columns([1, 1, 1])
            
            with col1:
                selected_file = st.selectbox("📁 Select File:", file_names, key="viewer_file")
            
            with col2:
                if selected_file:
                    workbook = st.session_state.excel_processor.workbooks[selected_file]
                    sheet_names = workbook.sheetnames
                    selected_sheet = st.selectbox("📋 Select Sheet:", sheet_names, key="viewer_sheet")
            
            with col3:
                if st.button("🔄 Refresh View", help="Refresh to see latest changes"):
                    st.rerun()
            
            # Show Excel viewer immediately
            if selected_file and selected_sheet:
                create_excel_like_viewer(selected_file, selected_sheet, st.session_state.excel_processor)
                
                # Show edit indicators
                if st.session_state.ai_instructor and st.session_state.ai_instructor.edit_history:
                    recent_edits = [edit for edit in st.session_state.ai_instructor.edit_history 
                                  if selected_file in edit['result'].get('location', '')]
                    if recent_edits:
                        st.info(f"🔥 {len(recent_edits)} recent edits made to this file!")
    
    # Show results AFTER the Excel viewer
    if st.session_state.analysis_result:
        st.divider()
        st.header("🎯 AI Analysis Results")
        
        # Display results with syntax highlighting
        st.markdown(st.session_state.analysis_result)
        
        # Show specific changes made
        if st.session_state.ai_instructor and st.session_state.ai_instructor.edit_history:
            st.subheader("✏️ Changes Made:")
            for edit in st.session_state.ai_instructor.edit_history[-5:]:  # Show last 5 edits
                st.success(f"✅ **{edit['result']['location']}**: '{edit['result'].get('old_value', '')}' → '{edit['result'].get('new_value', '')}'")
        
        # Download results
        st.download_button(
            "📥 Download Analysis Report",
            data=st.session_state.analysis_result,
            file_name="ai_excel_analysis.txt",
            mime="text/plain"
        )
        
        # Show edit history if available
        if st.session_state.ai_instructor and st.session_state.ai_instructor.edit_history:
            with st.expander("📝 Edit History"):
                for i, edit in enumerate(st.session_state.ai_instructor.edit_history[-10:], 1):
                    st.text(f"{i}. {edit['timestamp'][:19]}")
                    st.text(f"   Command: {edit['command']}")
                    st.text(f"   Result: {edit['result']['location']} - {edit['result'].get('old_value', '')} → {edit['result'].get('new_value', '')}")
                    st.divider()
    
    # Instructions and Examples
    if not st.session_state.files_loaded:
        st.divider()
        st.header("💡 Advanced Excel AI Features")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("🔧 Advanced Capabilities")
            st.markdown("""
            **Formula Support:**
            - Reads and preserves Excel formulas
            - Shows both formula and calculated values
            - Handles complex calculations correctly
            
            **Merged Cells:**
            - Detects and handles merged cell ranges
            - Maintains proper cell relationships
            - Preserves formatting structure
            
            **Smart Search:**
            - TF-IDF powered search across all data
            - Finds exact row and column locations
            - Supports complex queries and patterns
            
            **Precise Editing:**
            - Uses exact Excel cell references (A1, B5, etc.)
            - Maintains data types and formatting
            - Preserves formulas and relationships
            """)
        
        with col2:
            st.subheader("🎯 Supported Commands")
            st.markdown("""
            **Search Commands:**
            - `SEARCH: Total Liabilities` - Find specific text
            - `SEARCH: revenue data August` - Complex queries
            
            **Read Commands:**
            - `READ_CELL: file.xlsx Sheet1 A5` - Read specific cell
            - Shows both raw and calculated values
            
            **Edit Commands:**
            - `EDIT_CELL: file.xlsx Sheet1 A5 "New Value"`
            - `EDIT_CELL: file.xlsx Sheet1 B10 =SUM(B1:B9)`
            
            **AI understands natural language:**
            - "Change Total Liabilities text in Balance Sheet"
            - "Find the highest revenue value"
            - "Update all prices by 10%"
            """)
        
        st.subheader("📝 Example Natural Language Requests")
        
        examples = [
            "Find 'Total Liabilities & Shareholders' equity' and change it to 'Total Liabilities'",
            "Search for all revenue data and show me the totals",
            "Edit cell A32 in Balance Sheet to show 'Total Liabilities'",
            "Find all cells containing formulas and show their calculated values",
            "Replace 'Aug-22' with 'August 2022' in all headers",
            "Calculate the sum of column K from rows 15 to 25",
            "Find the cell with the highest value in the Cash column",
            "Update all percentage values by adding 5% to each"
        ]
        
        for i, example in enumerate(examples, 1):
            st.markdown(f"**{i}.** {example}")
        
        st.info("💡 **Tip:** The AI can handle complex Excel files with formulas, merged cells, and intricate formatting while maintaining data integrity!")

if __name__ == "__main__":
    main()
